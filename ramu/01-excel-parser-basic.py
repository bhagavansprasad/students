import openpyxl
print (openpyxl.__release__)
print (dir(openpyxl))

print(openpyxl.__major__)
print(openpyxl.__minor__)
print(openpyxl.__name__)
print(openpyxl.__package__)
print(openpyxl.__path__)

from openpyxl import load_workbook
#from openpyxl.style import Font, Fill, colors, PatternFill
from openpyxl.style import Font, Fill, Color, Fill
print (dir(openpyxl.style))

#import dump_cell

wb = load_workbook('/home/bhagavan/Downloads/revenue.xlsx')

#print worksheet names
print(wb.get_sheet_names())
for wsheet in wb.get_sheet_names():
    print(wsheet)

#opening worksheet in workbook
wsheet = wb.get_sheet_by_name('sales')

#max rows and cols
rcount =  wsheet.max_row
ccount =  wsheet.max_column
print("max rows filled :", rcount)
print("max cols filled :", ccount)

#print cell value
print("A2 :", wsheet['A2'].value) 
print("A3 :", wsheet['C3'].value) 

#print in row and col format
print("wsheet[1][0] :", wsheet[1][0].value) 
print("wsheet[2][1] :", wsheet[2][1].value) 
#print "wsheet[0][1] :", wsheet[0][1].value 

#print rows and columns
for row in range(1, rcount):
    flag = 1
    for col in range (0, ccount):
        if (wsheet[row][col].value != None):
            #print (str(wsheet[row][col].value)),
            print((wsheet[row][col].value), end == ' ')
            flag = 0

    print("")
    if (flag == 1):
       break;

print("")

#print rows and columns
for row in range(1, rcount):
    if not any(cell.value for cell in wsheet[row]):
        print("The %d row is empty" % row)
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None):
            print((str(wsheet[row][col].value)), end==' ')

    print("")

print("")

#print rows and columns even if there is any empty row in the middle
for row in range(1, rcount):
    for col in range (0, ccount):
        if (wsheet[row][col].value != None):
            print((str(wsheet[row][col].value)), end==' ')
    print("")

print("")

#Highlight cell, if the cell value is Vinay
#Change cell font size to 12, color Red, bold
for row in range(1, rcount+1):
    if not any(cell.value for cell in wsheet[row]):
        print("The %d row is empty" % row)
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None and wsheet[row][col].value == "Vinay" ):
            print((wsheet[row][col].font))
            wsheet[row][col].font = Font(size=12, bold=True, color=Color.RED)
            print("")

#Highlight row if the cell value is Ashish
#Change row font size to 12, color GREEN, bold
for row in range(1, rcount):
    if not any(cell.value for cell in wsheet[row]):
        print("The %d row is empty" % row)
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None and wsheet[row][col].value == "Ashish" ):
            for cell in wsheet[row:row]:
                cell.font = Font(size=12, bold=True, color=Color.GREEN)

#Highlight row background if the cell value is Kavitha
#Change row background to yellow
for row in range(1, rcount):
    if not any(cell.value for cell in wsheet[row]):
        print("The %d row is empty" % row)
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None and wsheet[row][col].value == "Kavitha" ):
            print((wsheet[row][col].fill))
            print((str(wsheet[row][col].value)))
            for cell in wsheet[row:row]:
                cell.fill = Fill(fill_type="solid", start_color=Color.YELLOW)

#Highlight column background in case cell value is > 70000
for row in range(1, rcount):
    if (wsheet[row][4].value != None and int(wsheet[row][4].value) > int(70000) ):
        print((str(wsheet[row][4].value)))
        cell = wsheet[row][4]
        cell.fill = Fill(fill_type="solid", start_color=Color.GREEN)

for word in dir(wsheet['C3']):
    print("print \"%-20s :\", cell.%s" %  (word, word))

wb.save('shared/revenue.xlsx')
#dump_cell.dump_cell_properties(wsheet[1][0])

exit(1)
